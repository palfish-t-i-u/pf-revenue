from __future__ import annotations

from datetime import datetime, timezone
from io import BytesIO
from pathlib import Path
import sys

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from payment_logic import GMV_CUTOFF, GMV_EXCHANGE_RATE
from payment_routes import (
    build_payment_export_workbook,
    get_payment_gmv_meta,
    parse_payment_import_file,
)


def test_parse_payment_import_file_reads_legacy_xlsx() -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SM Hanoi"
    ws.append(["header"] * 35)

    row = [None] * 35
    row[0] = "2026-05-30"
    row[3] = "Khach A"
    row[4] = "0900000000"
    row[5] = "123456"
    row[6] = "2/W- NEW 48 PHI+5"
    row[7] = "fixed"
    row[8] = "2026-05-30"
    row[10] = 3_700_000
    row[11] = 1000
    row[13] = "1st"
    row[14] = "Ads"
    row[21] = "Sale Legacy"
    row[33] = "In-house"
    ws.append(row)

    payload = BytesIO()
    wb.save(payload)

    parsed = parse_payment_import_file("legacy.xlsx", payload.getvalue())

    assert len(parsed) == 1
    assert parsed[0]["uid"] == "123456"
    assert parsed[0]["sale_name"] == "Sale Legacy"
    assert parsed[0]["team"] == "In-house"
    assert parsed[0]["real_pay_vnd"] == 3_700_000


def test_build_payment_export_workbook_writes_flattened_rows() -> None:
    payload = build_payment_export_workbook(
        [
            {
                "pay_time": "2026-06-09T08:30:00+00:00",
                "bank_day": "2026-06-09",
                "uid": "123456",
                "customers": {"full_name": "Khach A", "phone": "0900000000"},
                "sales": {"full_name": "Sale A", "short_code": "SA", "team": "In-house"},
                "channels": {"name": "Ads", "type": "Ads", "channel_code": "1133"},
                "packages": {"name": "Combo A", "fixed": "fixed"},
                "real_pay_vnd": 3_700_000,
                "gmv_rmb": 1000,
                "gmv_final": 1000,
                "payment_seq": "1st",
                "status": "active",
                "bank_matched": False,
                "crm_activated": False,
                "crm_order_id": None,
                "note": "hello",
                "payment_id": "pay-1",
            }
        ]
    )

    wb = openpyxl.load_workbook(BytesIO(payload))
    ws = wb.active

    assert ws.title == "Payments"
    assert ws["A1"].value == "Ngày thanh toán"
    assert ws["C2"].value == "123456"
    assert ws["D2"].value == "Khach A"
    assert ws["H2"].value == "Ads"
    assert ws["N2"].value == "active"


def test_get_payment_gmv_meta_matches_backend_rule() -> None:
    meta = get_payment_gmv_meta()

    assert meta["exchange_rate"] == float(GMV_EXCHANGE_RATE)
    assert meta["cutoff_at"] == GMV_CUTOFF.isoformat()
    assert datetime.fromisoformat(meta["cutoff_at"]).tzinfo == timezone.utc
