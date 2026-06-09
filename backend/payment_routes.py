"""Module Quản lý Doanh thu (payments) — song song Sổ doanh thu cũ."""

from __future__ import annotations

import csv
import io
from datetime import date, datetime, timezone
from decimal import Decimal
from typing import Any

import openpyxl
from fastapi import File, Header, HTTPException, Query, UploadFile
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field

from admin_routes import require_module_access, require_module_write
from payment_logic import compute_gmv_final, get_gmv_rule_meta
from rbac import require_min_role, resolve_actor
from sheet_row_parsers import parse_danang_row, parse_hcm_rev_row, parse_sm_hanoi_row

PAYMENT_SELECT = (
    "*, customers(full_name, phone), "
    "sales(short_code, full_name, team, khoi), "
    "channels(name, type, channel_code), "
    "packages(name, fixed)"
)
LEGACY_SHEET_PARSERS = {
    "hcm rev": parse_hcm_rev_row,
    "sm hanoi": parse_sm_hanoi_row,
    "danang rev": parse_danang_row,
}
EXPORT_HEADERS = [
    "Ngày thanh toán",
    "Ngày ngân hàng",
    "UID",
    "Khách hàng",
    "SĐT",
    "Sale",
    "Team",
    "Kênh",
    "Loại kênh",
    "Gói",
    "Tiền VNĐ",
    "GMV RMB",
    "GMV Final",
    "Trạng thái",
    "Lần TT",
    "Khớp NH",
    "Kích hoạt CRM",
    "CRM Order ID",
    "Note",
    "Payment ID",
]


def _match_legacy_sheet_parser(sheet_name: str):
    normalized = sheet_name.strip().lower()
    if normalized in LEGACY_SHEET_PARSERS:
        return LEGACY_SHEET_PARSERS[normalized]
    for key, parser in LEGACY_SHEET_PARSERS.items():
        if key in normalized or normalized in key:
            return parser
    return None


def _normalize_import_key(value: str) -> str:
    return value.strip().lower().replace(" ", "_").replace("-", "_")


def _first_present(row: dict[str, Any], *keys: str) -> Any:
    for key in keys:
        normalized = _normalize_import_key(key)
        if normalized in row and row[normalized] not in (None, ""):
            return row[normalized]
    return None


def _parse_normalized_import_row(row: dict[str, Any]) -> dict[str, Any] | None:
    uid_raw = _first_present(row, "uid")
    pay_time_raw = _first_present(row, "pay_time", "ngay_thanh_toan")
    real_pay_raw = _first_present(row, "real_pay_vnd", "tien_vnd", "vnd")

    if uid_raw in (None, "") and pay_time_raw in (None, "") and real_pay_raw in (None, ""):
        return None

    uid = str(uid_raw or "").strip()
    if not uid:
        raise HTTPException(400, "Thiếu uid trong file import")

    pay_time = _parse_pay_time(pay_time_raw)
    real_pay_vnd = float(Decimal(str(real_pay_raw or 0)))
    if real_pay_vnd <= 0:
        raise HTTPException(400, f"real_pay_vnd không hợp lệ cho uid {uid}")

    bank_day_raw = _first_present(row, "bank_day", "ngay_ngan_hang")
    bank_day = None
    if bank_day_raw not in (None, ""):
        parsed_day = _parse_pay_time(bank_day_raw)
        bank_day = parsed_day.date().isoformat()

    gmv_rmb_raw = _first_present(row, "gmv_rmb")
    gmv_rmb = float(Decimal(str(gmv_rmb_raw))) if gmv_rmb_raw not in (None, "") else None

    return {
        "uid": uid,
        "pay_time": pay_time,
        "bank_day": bank_day or pay_time.date().isoformat(),
        "real_pay_vnd": real_pay_vnd,
        "gmv_rmb": gmv_rmb,
        "team": str(_first_present(row, "team") or "").strip() or None,
        "sale_name": str(_first_present(row, "sale_name", "sale", "ten_sale") or "").strip() or None,
        "package_name": str(_first_present(row, "package_name", "package", "goi") or "").strip() or None,
        "fixed": str(_first_present(row, "fixed") or "").strip() or None,
        "channel_type_raw": str(_first_present(row, "channel_type", "channel_type_raw", "loai_kenh") or "").strip() or None,
        "channel_code": str(_first_present(row, "channel_code", "ma_kenh") or "").strip() or None,
        "channel_name": str(_first_present(row, "channel_name", "kenh") or "").strip() or None,
        "payment_seq": str(_first_present(row, "payment_seq", "lan_tt") or "").strip() or None,
        "note": str(_first_present(row, "note") or "").strip() or None,
        "customer_name": str(_first_present(row, "customer_name", "full_name", "khach_hang") or "").strip() or None,
        "customer_phone": str(_first_present(row, "customer_phone", "phone", "sdt") or "").strip() or None,
    }


def parse_payment_import_file(filename: str, content: bytes) -> list[dict[str, Any]]:
    suffix = filename.lower().rsplit(".", 1)[-1] if "." in filename else ""

    if suffix in {"xlsx", "xlsm"}:
        workbook = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        parsed_rows: list[dict[str, Any]] = []
        recognized_legacy_sheet = False

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            parser = _match_legacy_sheet_parser(sheet_name)
            if parser is not None:
                recognized_legacy_sheet = True
                values = list(sheet.iter_rows(values_only=True))
                for row_index, raw_row in enumerate(values[1:], start=2):
                    parsed = parser(list(raw_row))
                    if parsed:
                        parsed["_row"] = row_index
                        parsed["_sheet"] = sheet_name
                        parsed_rows.append(parsed)
                continue

            if recognized_legacy_sheet:
                continue

            rows = list(sheet.iter_rows(values_only=True))
            if not rows:
                continue
            headers = [_normalize_import_key(str(cell or "")) for cell in rows[0]]
            for row_index, raw_row in enumerate(rows[1:], start=2):
                normalized = {headers[idx]: raw_row[idx] for idx in range(min(len(headers), len(raw_row))) if headers[idx]}
                parsed = _parse_normalized_import_row(normalized)
                if parsed:
                    parsed["_row"] = row_index
                    parsed["_sheet"] = sheet_name
                    parsed_rows.append(parsed)
        workbook.close()
        return parsed_rows

    if suffix == "csv":
        decoded = content.decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(decoded))
        parsed_rows = []
        for row_index, row in enumerate(reader, start=2):
            normalized = {_normalize_import_key(key): value for key, value in row.items() if key}
            parsed = _parse_normalized_import_row(normalized)
            if parsed:
                parsed["_row"] = row_index
                parsed["_sheet"] = "csv"
                parsed_rows.append(parsed)
        return parsed_rows

    raise HTTPException(400, f"Định dạng file chưa hỗ trợ: .{suffix or 'unknown'}")


def get_payment_gmv_meta() -> dict[str, float | str]:
    return get_gmv_rule_meta()


def build_payment_export_workbook(items: list[dict[str, Any]]) -> bytes:
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Payments"
    worksheet.append(EXPORT_HEADERS)

    for item in items:
        customer = item.get("customers") or {}
        sale = item.get("sales") or {}
        channel = item.get("channels") or {}
        package = item.get("packages") or {}
        worksheet.append(
            [
                item.get("pay_time"),
                item.get("bank_day"),
                item.get("uid"),
                customer.get("full_name"),
                customer.get("phone"),
                sale.get("short_code") or sale.get("full_name"),
                item.get("team") or sale.get("team"),
                channel.get("name"),
                channel.get("type"),
                package.get("name"),
                item.get("real_pay_vnd"),
                item.get("gmv_rmb"),
                item.get("gmv_final"),
                item.get("status"),
                item.get("payment_seq"),
                item.get("bank_matched"),
                item.get("crm_activated"),
                item.get("crm_order_id"),
                item.get("note"),
                item.get("payment_id"),
            ]
        )

    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def _build_master_maps(sb) -> tuple[dict[str, dict[str, Any]], dict[str, str], dict[str, str]]:
    sales_res = sb.table("sales").select("id, full_name, short_code, team").execute()
    channels_res = sb.table("channels").select("id, name, type, channel_code").execute()
    packages_res = sb.table("packages").select("id, name").execute()

    sale_map: dict[str, dict[str, Any]] = {}
    for row in sales_res.data or []:
        full_name = str(row.get("full_name") or "").strip()
        short_code = str(row.get("short_code") or "").strip()
        if full_name:
            sale_map[full_name.lower()] = row
        if short_code:
            sale_map[short_code.lower()] = row

    channel_map: dict[str, str] = {}
    for row in channels_res.data or []:
        channel_id = row.get("id")
        name = str(row.get("name") or "").strip()
        channel_type = str(row.get("type") or "").strip()
        channel_code = str(row.get("channel_code") or "").strip()
        if name:
            channel_map[f"name|{name.lower()}"] = channel_id
        if channel_type:
            channel_map[f"type|{channel_type.lower()}"] = channel_id
        if channel_type or channel_code:
            channel_map[f"typecode|{channel_type.lower()}|{channel_code.lower()}"] = channel_id

    package_map = {
        str(row.get("name") or "").strip().lower(): row.get("id")
        for row in (packages_res.data or [])
        if str(row.get("name") or "").strip()
    }
    return sale_map, channel_map, package_map


def _prepare_import_payment_rows(sb, parsed_rows: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], list[dict[str, Any]], int]:
    sale_map, channel_map, package_map = _build_master_maps(sb)
    customer_rows: dict[str, dict[str, Any]] = {}
    payment_rows: list[dict[str, Any]] = []
    errors: list[dict[str, Any]] = []

    for index, row in enumerate(parsed_rows, start=1):
        sale_name = str(row.get("sale_name") or "").strip()
        sale = sale_map.get(sale_name.lower()) if sale_name else None
        if not sale:
            errors.append({"row": row.get("_row", index), "message": f"Không map được sale '{sale_name or 'trống'}'"})
            continue

        package_id = None
        package_name = str(row.get("package_name") or "").strip()
        if package_name:
            package_id = package_map.get(package_name.lower())
            if not package_id:
                errors.append({"row": row.get("_row", index), "message": f"Không tìm thấy package '{package_name}'"})
                continue

        channel_id = None
        channel_type = str(row.get("channel_type_raw") or "").strip()
        channel_code = str(row.get("channel_code") or "").strip()
        channel_name = str(row.get("channel_name") or "").strip()
        if channel_type or channel_code or channel_name:
            channel_id = (
                channel_map.get(f"typecode|{channel_type.lower()}|{channel_code.lower()}")
                or channel_map.get(f"name|{channel_name.lower()}")
                or channel_map.get(f"type|{channel_type.lower()}")
            )
            if not channel_id:
                errors.append({"row": row.get("_row", index), "message": f"Không tìm thấy channel cho '{channel_name or channel_type or channel_code}'"})
                continue

        uid = str(row.get("uid") or "").strip()
        pay_time = _parse_pay_time(row.get("pay_time"))
        real_vnd = Decimal(str(row.get("real_pay_vnd") or 0))
        gmv_rmb_raw = row.get("gmv_rmb")
        gmv_rmb = Decimal(str(gmv_rmb_raw)) if gmv_rmb_raw not in (None, "") else None
        gmv_final = compute_gmv_final(pay_time, real_vnd, gmv_rmb)

        customer_rows[uid] = {
            "uid": uid,
            "full_name": str(row.get("customer_name") or "").strip() or None,
            "phone": str(row.get("customer_phone") or "").strip() or None,
            "first_seen": row.get("bank_day") or pay_time.date().isoformat(),
        }
        payment_rows.append(
            {
                "uid": uid,
                "pay_time": pay_time.isoformat(),
                "bank_day": row.get("bank_day") or pay_time.date().isoformat(),
                "package_id": package_id,
                "sale_id": sale["id"],
                "channel_id": channel_id,
                "real_pay_vnd": float(real_vnd),
                "gmv_rmb": float(gmv_rmb) if gmv_rmb is not None else None,
                "gmv_final": float(gmv_final),
                "payment_seq": str(row.get("payment_seq") or "").strip() or None,
                "team": str(sale.get("team") or row.get("team") or "").strip(),
                "note": str(row.get("note") or "").strip() or None,
                "status": "active",
            }
        )

    if customer_rows:
        sb.table("customers").upsert(list(customer_rows.values()), on_conflict="uid").execute()

    return payment_rows, errors, len(customer_rows)

def _parse_pay_time(value: Any) -> datetime:
    if isinstance(value, datetime):
        dt = value
    elif isinstance(value, str):
        dt = datetime.fromisoformat(value.replace("Z", "+00:00"))
    else:
        raise HTTPException(400, "pay_time không hợp lệ")
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    return dt


def _now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def _apply_payment_filters(
    q,
    *,
    search: str = "",
    date_from: date | None = None,
    date_to: date | None = None,
    team: str = "",
    channel_id: str = "",
    sale_id: str = "",
    package_id: str = "",
    status: str = "",
    bank_matched: str = "",
    crm_activated: str = "",
    sb=None,
):
    q = q.is_("deleted_at", "null")
    if search.strip():
        term = search.strip()
        parts = [
            f"uid.ilike.%{term}%",
            f"note.ilike.%{term}%",
            f"payment_seq.ilike.%{term}%",
            f"team.ilike.%{term}%",
        ]
        if sb is not None:
            extra_uids = _search_customer_uids(sb, term)
            if extra_uids:
                parts.append(f"uid.in.({','.join(extra_uids)})")
            sale_ids = _search_sale_ids(sb, term)
            if sale_ids:
                parts.append(f"sale_id.in.({','.join(sale_ids)})")
            channel_ids = _search_channel_ids(sb, term)
            if channel_ids:
                parts.append(f"channel_id.in.({','.join(channel_ids)})")
            package_ids = _search_package_ids(sb, term)
            if package_ids:
                parts.append(f"package_id.in.({','.join(package_ids)})")
        q = q.or_(",".join(parts))
    if date_from:
        q = q.gte("pay_time", f"{date_from.isoformat()}T00:00:00")
    if date_to:
        q = q.lte("pay_time", f"{date_to.isoformat()}T23:59:59")
    if team:
        q = q.eq("team", team.strip())
    if channel_id:
        q = q.eq("channel_id", channel_id.strip())
    if sale_id:
        q = q.eq("sale_id", sale_id.strip())
    if package_id:
        q = q.eq("package_id", package_id.strip())
    if status:
        q = q.eq("status", status.strip())
    if bank_matched in ("true", "false"):
        q = q.eq("bank_matched", bank_matched == "true")
    if crm_activated in ("true", "false"):
        q = q.eq("crm_activated", crm_activated == "true")
    return q



def _fetch_payment_row(sb, payment_id: str) -> dict[str, Any]:
    res = (
        sb.table("payments")
        .select(PAYMENT_SELECT)
        .eq("payment_id", payment_id)
        .is_("deleted_at", "null")
        .limit(1)
        .execute()
    )
    if not res.data:
        raise HTTPException(404, f"Payment {payment_id} không tồn tại")
    return res.data[0]


def _search_customer_uids(sb, term: str) -> list[str]:
    res = (
        sb.table("customers")
        .select("uid")
        .or_(f"uid.ilike.%{term}%,full_name.ilike.%{term}%,phone.ilike.%{term}%")
        .limit(50)
        .execute()
    )
    return [str(r["uid"]) for r in (res.data or []) if r.get("uid")]


def _search_sale_ids(sb, term: str) -> list[str]:
    res = (
        sb.table("sales")
        .select("id")
        .or_(f"full_name.ilike.%{term}%,short_code.ilike.%{term}%")
        .limit(20)
        .execute()
    )
    return [str(r["id"]) for r in (res.data or []) if r.get("id")]


def _search_channel_ids(sb, term: str) -> list[str]:
    res = (
        sb.table("channels")
        .select("id")
        .or_(f"name.ilike.%{term}%")
        .limit(20)
        .execute()
    )
    return [str(r["id"]) for r in (res.data or []) if r.get("id")]


def _search_package_ids(sb, term: str) -> list[str]:
    res = (
        sb.table("packages")
        .select("id")
        .or_(f"name.ilike.%{term}%")
        .limit(20)
        .execute()
    )
    return [str(r["id"]) for r in (res.data or []) if r.get("id")]


class PaymentCreate(BaseModel):
    uid: str
    pay_time: datetime
    package_id: str | None = None
    sale_id: str
    channel_id: str | None = None
    real_pay_vnd: float = Field(..., gt=0)
    gmv_rmb: float | None = None
    payment_seq: str | None = None
    note: str | None = None
    bank_day: date | None = None
    customer_name: str | None = None
    customer_phone: str | None = None


class PaymentPatch(BaseModel):
    uid: str | None = None
    pay_time: datetime | None = None
    real_pay_vnd: float | None = None
    gmv_rmb: float | None = None
    sale_id: str | None = None
    channel_id: str | None = None
    package_id: str | None = None
    payment_seq: str | None = None
    note: str | None = None
    bank_day: date | None = None
    team: str | None = None


class LinkCrmBody(BaseModel):
    crm_order_id: str = Field(..., min_length=1)


class SaleCreate(BaseModel):
    full_name: str
    short_code: str | None = None
    team: str | None = None
    khoi: str | None = None
    active: bool = True


class SalePatch(BaseModel):
    full_name: str | None = None
    short_code: str | None = None
    team: str | None = None
    khoi: str | None = None
    active: bool | None = None


class ChannelCreate(BaseModel):
    channel_code: str | None = None
    name: str | None = None
    type: str | None = None


class ChannelPatch(BaseModel):
    channel_code: str | None = None
    name: str | None = None
    type: str | None = None


class PackageCreate(BaseModel):
    name: str
    fixed: str | None = None


class PackagePatch(BaseModel):
    name: str | None = None
    fixed: str | None = None


class CustomerPatch(BaseModel):
    full_name: str | None = None
    phone: str | None = None


def _trim(value: str | None) -> str | None:
    if value is None:
        return None
    s = value.strip()
    return s or None


def _trim_patch(body: BaseModel) -> dict[str, Any]:
    update: dict[str, Any] = {}
    for key, value in body.model_dump(exclude_none=True).items():
        if isinstance(value, str):
            trimmed = _trim(value)
            if trimmed is not None:
                update[key] = trimmed
        else:
            update[key] = value
    return update


def _patch_master_row(
    sb,
    *,
    table: str,
    id_field: str,
    id_value: str,
    update: dict[str, Any],
    select: str,
) -> dict[str, Any]:
    if not update:
        raise HTTPException(400, "Không có dữ liệu cập nhật")
    res = (
        sb.table(table)
        .update(update)
        .eq(id_field, id_value)
        .select(select)
        .execute()
    )
    if not res.data:
        raise HTTPException(404, f"{table} {id_value} không tồn tại")
    return res.data[0]


def register_payment_routes(app, sb_getter) -> None:
    def _sb():
        sb = sb_getter()
        if not sb:
            raise HTTPException(503, "Supabase chưa cấu hình")
        return sb

    @app.get("/api/v1/payments/meta", tags=["Payments"])
    def get_payment_meta(authorization: str | None = Header(None)):
        sb = _sb()
        actor = resolve_actor(sb, authorization)
        require_module_access(sb, actor, "payments")
        return {"gmv_rule": get_payment_gmv_meta()}

    @app.get("/api/v1/payments", tags=["Payments"])
    def list_payments(
        search: str = Query(""),
        date_from: date | None = Query(None, alias="from"),
        date_to: date | None = Query(None, alias="to"),
        team: str = Query(""),
        channel_id: str = Query(""),
        sale_id: str = Query(""),
        package_id: str = Query(""),
        status: str = Query(""),
        bank_matched: str = Query(""),
        crm_activated: str = Query(""),
        page: int = Query(1, ge=1),
        page_size: int = Query(50, ge=1, le=200),
        authorization: str | None = Header(None),
    ):
        sb = _sb()
        actor = resolve_actor(sb, authorization)
        require_module_access(sb, actor, "payments")

        offset = (page - 1) * page_size
        q = sb.table("payments").select(PAYMENT_SELECT, count="exact")
        q = _apply_payment_filters(
            q,
            search=search,
            date_from=date_from,
            date_to=date_to,
            team=team,
            channel_id=channel_id,
            sale_id=sale_id,
            package_id=package_id,
            status=status,
            bank_matched=bank_matched,
            crm_activated=crm_activated,
            sb=sb,
        )

        q = q.order("pay_time", desc=True).range(offset, offset + page_size - 1)
        res = q.execute()
        items = res.data or []
        total = int(res.count or len(items))

        rpc_params: dict[str, Any] = {}
        if search.strip():
            rpc_params["p_search"] = search.strip()
            if sb is not None:
                extra_uids = _search_customer_uids(sb, search.strip())
                if extra_uids:
                    rpc_params["p_extra_uids"] = extra_uids
        if date_from:
            rpc_params["p_date_from"] = f"{date_from.isoformat()}T00:00:00+00:00"
        if date_to:
            rpc_params["p_date_to"] = f"{date_to.isoformat()}T23:59:59+00:00"
        if team:
            rpc_params["p_team"] = team.strip()
        if channel_id:
            rpc_params["p_channel_id"] = channel_id.strip()
        if sale_id:
            rpc_params["p_sale_id"] = sale_id.strip()
        if status:
            rpc_params["p_status"] = status.strip()
        if bank_matched:
            rpc_params["p_bank_matched"] = bank_matched
        if crm_activated:
            rpc_params["p_crm_activated"] = crm_activated

        summary_res = sb.rpc("payments_summary", rpc_params).execute()
        summary = summary_res.data or {
            "count": 0, "gmv_final": 0, "real_pay_vnd": 0,
            "unmatched_bank": 0, "uncrm": 0,
        }

        return {"items": items, "total": total, "summary": summary}

    @app.get("/api/v1/payments/export", tags=["Payments"])
    def export_payments(
        search: str = Query(""),
        date_from: date | None = Query(None, alias="from"),
        date_to: date | None = Query(None, alias="to"),
        team: str = Query(""),
        channel_id: str = Query(""),
        sale_id: str = Query(""),
        package_id: str = Query(""),
        status: str = Query(""),
        bank_matched: str = Query(""),
        crm_activated: str = Query(""),
        authorization: str | None = Header(None),
    ):
        sb = _sb()
        actor = resolve_actor(sb, authorization)
        require_module_access(sb, actor, "payments")

        query = sb.table("payments").select(PAYMENT_SELECT)
        query = _apply_payment_filters(
            query,
            search=search,
            date_from=date_from,
            date_to=date_to,
            team=team,
            channel_id=channel_id,
            sale_id=sale_id,
            package_id=package_id,
            status=status,
            bank_matched=bank_matched,
            crm_activated=crm_activated,
            sb=sb,
        )
        items = query.order("pay_time", desc=True).limit(5000).execute().data or []
        workbook_bytes = build_payment_export_workbook(items)
        filename = f"payments_{datetime.now(timezone.utc).date().isoformat()}.xlsx"

        return StreamingResponse(
            io.BytesIO(workbook_bytes),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )

    @app.post("/api/v1/payments/import", tags=["Payments"])
    async def import_payments(
        file: UploadFile = File(...),
        authorization: str | None = Header(None),
    ):
        sb = _sb()
        actor = resolve_actor(sb, authorization)
        require_module_write(sb, actor, "payments")

        payload = await file.read()
        parsed_rows = parse_payment_import_file(file.filename or "upload", payload)
        if not parsed_rows:
            return {"inserted": 0, "skipped": 0, "errors": [{"row": 0, "message": "Không đọc được dòng hợp lệ nào"}]}

        payment_rows, errors, _customer_count = _prepare_import_payment_rows(sb, parsed_rows)
        inserted = 0
        skipped = 0

        for row, source in zip(payment_rows, [r for r in parsed_rows if not any(e["row"] == r.get("_row") for e in errors)], strict=False):
            try:
                sb.table("payments").insert(row).execute()
                inserted += 1
            except Exception as exc:
                skipped += 1
                message = "Trùng uid + pay_time + real_pay_vnd" if "duplicate" in str(exc).lower() or "payments_bizkey" in str(exc).lower() else str(exc)
                errors.append({"row": source.get("_row"), "message": message})

        return {"inserted": inserted, "skipped": skipped, "errors": sorted(errors, key=lambda x: x["row"])}

    @app.post("/api/v1/payments", tags=["Payments"])
    def create_payment(
        body: PaymentCreate,
        authorization: str | None = Header(None),
    ):
        sb = _sb()
        actor = resolve_actor(sb, authorization)
        require_module_write(sb, actor, "payments")

        uid = body.uid.strip()
        if not uid:
            raise HTTPException(400, "uid không được rỗng")

        existing = sb.table("customers").select("uid").eq("uid", uid).limit(1).execute()
        if not existing.data:
            sb.table("customers").insert(
                {
                    "uid": uid,
                    "full_name": (body.customer_name or "").strip() or None,
                    "phone": (body.customer_phone or "").strip() or None,
                    "first_seen": body.pay_time.date().isoformat(),
                }
            ).execute()

        sale_res = (
            sb.table("sales")
            .select("team")
            .eq("id", body.sale_id)
            .limit(1)
            .execute()
        )
        if not sale_res.data:
            raise HTTPException(400, "sale_id không tồn tại")
        team_raw = str(sale_res.data[0].get("team") or "").strip()
        if not team_raw:
            raise HTTPException(400, "sale chưa có team")

        pay_time = _parse_pay_time(body.pay_time)
        real_vnd = Decimal(str(body.real_pay_vnd))
        gmv_rmb = Decimal(str(body.gmv_rmb)) if body.gmv_rmb is not None else None
        gmv_final = compute_gmv_final(pay_time, real_vnd, gmv_rmb)

        row = {
            "uid": uid,
            "pay_time": pay_time.isoformat(),
            "bank_day": (body.bank_day or pay_time.date()).isoformat(),
            "package_id": body.package_id,
            "sale_id": body.sale_id,
            "channel_id": body.channel_id,
            "real_pay_vnd": float(real_vnd),
            "gmv_rmb": float(gmv_rmb) if gmv_rmb is not None else None,
            "gmv_final": float(gmv_final),
            "payment_seq": body.payment_seq,
            "team": team_raw,
            "note": body.note,
            "status": "active",
        }
        try:
            res = sb.table("payments").insert(row).execute()
        except Exception as exc:
            if "payments_bizkey" in str(exc).lower() or "duplicate" in str(exc).lower():
                raise HTTPException(409, "Trùng uid + pay_time + real_pay_vnd") from exc
            raise HTTPException(500, f"Không tạo payment: {exc}") from exc
        if not res.data:
            raise HTTPException(500, "Không tạo payment")
        return _fetch_payment_row(sb, res.data[0]["payment_id"])

    @app.patch("/api/v1/payments/{payment_id}", tags=["Payments"])
    def patch_payment(
        payment_id: str,
        body: PaymentPatch,
        authorization: str | None = Header(None),
    ):
        sb = _sb()
        actor = resolve_actor(sb, authorization)
        require_module_write(sb, actor, "payments")

        update = body.model_dump(exclude_none=True)
        if not update:
            raise HTTPException(400, "Không có dữ liệu cập nhật")

        current = (
            sb.table("payments")
            .select("pay_time, real_pay_vnd, gmv_rmb, sale_id, team")
            .eq("payment_id", payment_id)
            .is_("deleted_at", "null")
            .limit(1)
            .execute()
        )
        if not current.data:
            raise HTTPException(404, f"Payment {payment_id} không tồn tại")
        cur = current.data[0]

        if any(k in update for k in ("pay_time", "real_pay_vnd", "gmv_rmb")):
            pt_raw = update.get("pay_time", cur["pay_time"])
            pt = _parse_pay_time(pt_raw)
            rpv = Decimal(str(update.get("real_pay_vnd", cur["real_pay_vnd"])))
            grmb_raw = update.get("gmv_rmb", cur.get("gmv_rmb"))
            grmb = Decimal(str(grmb_raw)) if grmb_raw is not None else None
            update["gmv_final"] = float(compute_gmv_final(pt, rpv, grmb))
            if "pay_time" in update:
                update["pay_time"] = pt.isoformat()

        if "sale_id" in update and "team" not in update:
            sale = (
                sb.table("sales")
                .select("team")
                .eq("id", update["sale_id"])
                .limit(1)
                .execute()
            )
            if not sale.data:
                raise HTTPException(400, "sale_id không tồn tại")
            update["team"] = sale.data[0].get("team") or cur.get("team") or ""

        if "bank_day" in update and isinstance(update["bank_day"], date):
            update["bank_day"] = update["bank_day"].isoformat()

        update["updated_at"] = _now_iso()
        res = (
            sb.table("payments")
            .update(update)
            .eq("payment_id", payment_id)
            .is_("deleted_at", "null")
            .execute()
        )
        if not res.data:
            return _fetch_payment_row(sb, payment_id)
        return _fetch_payment_row(sb, payment_id)

    @app.post("/api/v1/payments/{payment_id}/refund", tags=["Payments"])
    def refund_payment(payment_id: str, authorization: str | None = Header(None)):
        sb = _sb()
        actor = resolve_actor(sb, authorization)
        require_module_write(sb, actor, "payments")
        sb.table("payments").update(
            {"status": "refunded", "updated_at": _now_iso()}
        ).eq("payment_id", payment_id).is_("deleted_at", "null").execute()
        return {"ok": True, "status": "refunded", "payment_id": payment_id}

    @app.post("/api/v1/payments/{payment_id}/restore", tags=["Payments"])
    def restore_payment(payment_id: str, authorization: str | None = Header(None)):
        sb = _sb()
        actor = resolve_actor(sb, authorization)
        require_module_write(sb, actor, "payments")
        sb.table("payments").update(
            {"status": "active", "updated_at": _now_iso()}
        ).eq("payment_id", payment_id).is_("deleted_at", "null").execute()
        return {"ok": True, "status": "active", "payment_id": payment_id}

    @app.post("/api/v1/payments/{payment_id}/link-crm", tags=["Payments"])
    def link_crm(
        payment_id: str,
        body: LinkCrmBody,
        authorization: str | None = Header(None),
    ):
        sb = _sb()
        actor = resolve_actor(sb, authorization)
        require_module_write(sb, actor, "payments")
        sb.table("payments").update(
            {
                "crm_order_id": body.crm_order_id.strip(),
                "crm_activated": True,
                "activated_at": date.today().isoformat(),
                "updated_at": _now_iso(),
            }
        ).eq("payment_id", payment_id).is_("deleted_at", "null").execute()
        return {"ok": True, "payment_id": payment_id, "crm_order_id": body.crm_order_id}

    @app.delete("/api/v1/payments/{payment_id}", tags=["Payments"])
    def delete_payment(payment_id: str, authorization: str | None = Header(None)):
        sb = _sb()
        actor = resolve_actor(sb, authorization)
        require_min_role(actor, "system")
        sb.table("payments").update({"deleted_at": _now_iso()}).eq(
            "payment_id", payment_id
        ).execute()
        return {"ok": True, "payment_id": payment_id}

    @app.get("/api/v1/customers/search", tags=["Payments"])
    def search_customers(
        q: str = Query(..., min_length=1),
        authorization: str | None = Header(None),
    ):
        sb = _sb()
        actor = resolve_actor(sb, authorization)
        require_module_access(sb, actor, "payments")
        term = q.strip()
        res = (
            sb.table("customers")
            .select("uid, full_name, phone")
            .or_(f"uid.ilike.%{term}%,full_name.ilike.%{term}%,phone.ilike.%{term}%")
            .limit(20)
            .execute()
        )
        return res.data or []

    @app.get("/api/v1/payments/master/sales", tags=["Payments"])
    def list_sales_master(authorization: str | None = Header(None)):
        sb = _sb()
        actor = resolve_actor(sb, authorization)
        require_module_access(sb, actor, "payments")
        res = (
            sb.table("sales")
            .select("id, full_name, short_code, team, khoi, active")
            .eq("active", True)
            .order("full_name")
            .execute()
        )
        return res.data or []

    @app.get("/api/v1/payments/master/channels", tags=["Payments"])
    def list_channels_master(authorization: str | None = Header(None)):
        sb = _sb()
        actor = resolve_actor(sb, authorization)
        require_module_access(sb, actor, "payments")
        res = (
            sb.table("channels")
            .select("id, channel_code, name, type")
            .order("type")
            .execute()
        )
        return res.data or []

    @app.get("/api/v1/payments/master/packages", tags=["Payments"])
    def list_packages_master(authorization: str | None = Header(None)):
        sb = _sb()
        actor = resolve_actor(sb, authorization)
        require_module_access(sb, actor, "payments")
        res = sb.table("packages").select("id, name, fixed").order("name").execute()
        return res.data or []

    @app.post("/api/v1/payments/master/sales", tags=["Payments"])
    def create_sale_master(
        body: SaleCreate,
        authorization: str | None = Header(None),
    ):
        sb = _sb()
        actor = resolve_actor(sb, authorization)
        require_module_write(sb, actor, "payments")
        full_name = _trim(body.full_name)
        if not full_name:
            raise HTTPException(400, "full_name không được rỗng")
        row = {
            "full_name": full_name,
            "short_code": _trim(body.short_code),
            "team": _trim(body.team),
            "khoi": _trim(body.khoi),
            "active": body.active,
        }
        res = sb.table("sales").insert(row).select(
            "id, full_name, short_code, team, khoi, active"
        ).execute()
        if not res.data:
            raise HTTPException(500, "Không tạo sale")
        return res.data[0]

    @app.patch("/api/v1/payments/master/sales/{sale_id}", tags=["Payments"])
    def patch_sale_master(
        sale_id: str,
        body: SalePatch,
        authorization: str | None = Header(None),
    ):
        sb = _sb()
        actor = resolve_actor(sb, authorization)
        require_module_write(sb, actor, "payments")
        update = _trim_patch(body)
        if "full_name" in update and not update["full_name"]:
            raise HTTPException(400, "full_name không được rỗng")
        return _patch_master_row(
            sb,
            table="sales",
            id_field="id",
            id_value=sale_id,
            update=update,
            select="id, full_name, short_code, team, khoi, active",
        )

    @app.post("/api/v1/payments/master/channels", tags=["Payments"])
    def create_channel_master(
        body: ChannelCreate,
        authorization: str | None = Header(None),
    ):
        sb = _sb()
        actor = resolve_actor(sb, authorization)
        require_module_write(sb, actor, "payments")
        channel_code = _trim(body.channel_code)
        name = _trim(body.name)
        ch_type = _trim(body.type)
        if not name and not ch_type:
            name = channel_code or "Other"
        row = {
            "channel_code": channel_code,
            "name": name,
            "type": ch_type,
        }
        res = sb.table("channels").insert(row).select(
            "id, channel_code, name, type"
        ).execute()
        if not res.data:
            raise HTTPException(500, "Không tạo channel")
        return res.data[0]

    @app.patch("/api/v1/payments/master/channels/{channel_id}", tags=["Payments"])
    def patch_channel_master(
        channel_id: str,
        body: ChannelPatch,
        authorization: str | None = Header(None),
    ):
        sb = _sb()
        actor = resolve_actor(sb, authorization)
        require_module_write(sb, actor, "payments")
        return _patch_master_row(
            sb,
            table="channels",
            id_field="id",
            id_value=channel_id,
            update=_trim_patch(body),
            select="id, channel_code, name, type",
        )

    @app.post("/api/v1/payments/master/packages", tags=["Payments"])
    def create_package_master(
        body: PackageCreate,
        authorization: str | None = Header(None),
    ):
        sb = _sb()
        actor = resolve_actor(sb, authorization)
        require_module_write(sb, actor, "payments")
        name = _trim(body.name)
        if not name:
            raise HTTPException(400, "name không được rỗng")
        row = {"name": name, "fixed": _trim(body.fixed)}
        res = sb.table("packages").insert(row).select("id, name, fixed").execute()
        if not res.data:
            raise HTTPException(500, "Không tạo package")
        return res.data[0]

    @app.patch("/api/v1/payments/master/packages/{package_id}", tags=["Payments"])
    def patch_package_master(
        package_id: str,
        body: PackagePatch,
        authorization: str | None = Header(None),
    ):
        sb = _sb()
        actor = resolve_actor(sb, authorization)
        require_module_write(sb, actor, "payments")
        update = _trim_patch(body)
        if "name" in update and not update["name"]:
            raise HTTPException(400, "name không được rỗng")
        return _patch_master_row(
            sb,
            table="packages",
            id_field="id",
            id_value=package_id,
            update=update,
            select="id, name, fixed",
        )

    @app.patch("/api/v1/payments/master/customers/{uid}", tags=["Payments"])
    def patch_customer_master(
        uid: str,
        body: CustomerPatch,
        authorization: str | None = Header(None),
    ):
        sb = _sb()
        actor = resolve_actor(sb, authorization)
        require_module_write(sb, actor, "payments")
        customer_uid = uid.strip()
        if not customer_uid:
            raise HTTPException(400, "uid không được rỗng")
        return _patch_master_row(
            sb,
            table="customers",
            id_field="uid",
            id_value=customer_uid,
            update=_trim_patch(body),
            select="uid, full_name, phone, first_seen, created_at",
        )
