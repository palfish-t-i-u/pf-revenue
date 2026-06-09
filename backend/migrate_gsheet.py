"""One-time migration: Google Sheet (.xlsx) → Supabase tables.

Usage:
    python migrate_gsheet.py "path/to/All File Thu Hien.xlsx"            # dry-run (mặc định)
    python migrate_gsheet.py "path/to/All File Thu Hien.xlsx" --commit   # chạy thật

Dry-run: parse + validate + báo cáo thống kê, KHÔNG ghi DB.
--commit: parse + validate + ghi DB.

Requires: SUPABASE_URL + SUPABASE_SERVICE_ROLE_KEY in .env or environment.
"""

from __future__ import annotations

import os
import sys
from collections import defaultdict
from datetime import datetime, timezone
from decimal import Decimal
from pathlib import Path

import openpyxl
from dotenv import load_dotenv
from supabase import create_client

from payment_logic import compute_gmv_final
from sheet_row_parsers import (
    DEFAULT_TY_GIA,
    parse_danang_row,
    parse_hcm_rev_row,
    parse_sm_hanoi_row,
)

load_dotenv(Path(__file__).parent / ".env")

# ── Sheet tab → parser mapping ──────────────────────────────────
# Chỉ 3 sheet chứa dữ liệu gốc. Các sheet khác (HN Inhouse 1/2,
# HN Offline, HCM, GMV, note, ALL SUM) là báo cáo tổng hợp — bỏ qua.
SHEET_PARSERS = {
    "hcm rev": parse_hcm_rev_row,
    "sm hanoi": parse_sm_hanoi_row,
    "danang rev": parse_danang_row,
}

SKIP_ROWS = 1  # Skip header row(s)

# ── Validation thresholds ───────────────────────────────────────
VND_MIN = 50_000           # < 50K VND → cảnh báo (có thể nhập nhầm đơn vị)
VND_MAX = 2_000_000_000    # > 2 tỷ VND → cảnh báo
GMV_RATIO_LOW = 0.05       # GMV/expected < 5% → cảnh báo (có thể nhầm tiền tệ)
GMV_RATIO_HIGH = 20.0      # GMV/expected > 20x → cảnh báo


def match_parser(tab_name: str):
    """Find parser by fuzzy-matching tab name."""
    normalized = tab_name.strip().lower()
    if normalized in SHEET_PARSERS:
        return SHEET_PARSERS[normalized]
    for key, parser in SHEET_PARSERS.items():
        if key in normalized or normalized in key:
            return parser
    return None


def read_all_sheets(xlsx_path: str) -> list[dict]:
    """Read xlsx and parse all recognized sheets."""
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    all_parsed = []

    print(f"\n📄 File: {xlsx_path}")
    print(f"   Sheets: {wb.sheetnames}\n")

    for sheet_name in wb.sheetnames:
        parser = match_parser(sheet_name)
        if not parser:
            print(f"   ⏭️  Skip sheet '{sheet_name}' — no parser matched")
            continue

        ws = wb[sheet_name]
        rows_raw = list(ws.iter_rows(values_only=True))
        data_rows = rows_raw[SKIP_ROWS:]

        parsed_count = 0
        for row_idx, row in enumerate(data_rows, start=SKIP_ROWS + 1):
            row_list = list(row)
            result = parser(row_list)
            if result:
                result["_sheet"] = sheet_name
                result["_row"] = row_idx
                all_parsed.append(result)
                parsed_count += 1

        print(f"   ✅ Sheet '{sheet_name}' → {parser.__name__} → {parsed_count} rows parsed")

    wb.close()
    print(f"\n   Total parsed: {len(all_parsed)} rows")
    return all_parsed


# ── Validation ──────────────────────────────────────────────────

def validate_rows(parsed_rows: list[dict]) -> list[dict]:
    """Validate parsed rows. Returns list of warnings (not blocking)."""
    warnings = []

    seen_bizkeys: dict[str, list] = defaultdict(list)
    vnd_values = []
    team_counts = defaultdict(int)
    sale_counts = defaultdict(int)

    for i, row in enumerate(parsed_rows):
        vnd = row.get("real_pay_vnd", 0)
        gmv = row.get("gmv_rmb", 0) or 0
        uid = row.get("uid", "?")
        sheet = row.get("_sheet", "?")
        row_num = row.get("_row", "?")
        loc = f"[{sheet} row {row_num}]"

        # ── VND range check ──
        if vnd < VND_MIN:
            warnings.append({
                "type": "VND_TOO_LOW",
                "severity": "high",
                "detail": f"{loc} uid={uid} VND={vnd:,.0f} — có thể thiếu 000?",
            })
        elif vnd > VND_MAX:
            warnings.append({
                "type": "VND_TOO_HIGH",
                "severity": "high",
                "detail": f"{loc} uid={uid} VND={vnd:,.0f} — có thể nhập thừa số 0?",
            })

        # ── GMV vs VND ratio check ──
        if vnd > 0 and gmv > 0:
            expected_gmv = float(vnd) / float(DEFAULT_TY_GIA)
            if expected_gmv > 0:
                ratio = gmv / expected_gmv
                if ratio < GMV_RATIO_LOW:
                    warnings.append({
                        "type": "GMV_TOO_LOW",
                        "severity": "medium",
                        "detail": f"{loc} uid={uid} VND={vnd:,.0f} GMV={gmv:.2f} "
                                  f"ratio={ratio:.2f} — GMV quá nhỏ so với VND",
                    })
                elif ratio > GMV_RATIO_HIGH:
                    warnings.append({
                        "type": "GMV_TOO_HIGH",
                        "severity": "medium",
                        "detail": f"{loc} uid={uid} VND={vnd:,.0f} GMV={gmv:.2f} "
                                  f"ratio={ratio:.2f} — GMV quá lớn so với VND",
                    })

        # ── Duplicate detection ──
        pay_time = row.get("pay_time")
        pt_str = pay_time.isoformat() if hasattr(pay_time, "isoformat") else str(pay_time)
        bizkey = f"{uid}|{pt_str}|{vnd}"
        seen_bizkeys[bizkey].append(loc)

        # ── Stats ──
        vnd_values.append(vnd)
        team_counts[row.get("team", "?")] += 1
        sale_counts[row.get("sale_name", "?")] += 1

    # ── Duplicate warnings ──
    dupe_count = 0
    for bizkey, locations in seen_bizkeys.items():
        if len(locations) > 1:
            dupe_count += len(locations) - 1
            if len(locations) <= 5:  # Only log first few
                warnings.append({
                    "type": "DUPLICATE",
                    "severity": "low",
                    "detail": f"bizkey={bizkey} xuất hiện {len(locations)} lần: {locations}",
                })

    # ── Summary stats ──
    print(f"\n{'='*60}")
    print(f"📊 VALIDATION REPORT")
    print(f"{'='*60}")

    if vnd_values:
        print(f"\n💰 VND distribution:")
        print(f"   Min:    {min(vnd_values):>15,.0f}")
        print(f"   Max:    {max(vnd_values):>15,.0f}")
        print(f"   Avg:    {sum(vnd_values)/len(vnd_values):>15,.0f}")
        print(f"   Total:  {sum(vnd_values):>15,.0f}")

    print(f"\n👥 By team:")
    for team, count in sorted(team_counts.items(), key=lambda x: -x[1]):
        print(f"   {team:<20} {count:>6} rows")

    print(f"\n👤 Top 10 sales:")
    for name, count in sorted(sale_counts.items(), key=lambda x: -x[1])[:10]:
        print(f"   {name:<25} {count:>6} rows")

    # ── Warnings summary ──
    high = [w for w in warnings if w["severity"] == "high"]
    medium = [w for w in warnings if w["severity"] == "medium"]
    low = [w for w in warnings if w["severity"] == "low"]

    print(f"\n⚠️  Warnings: {len(high)} high, {len(medium)} medium, {len(low)} low")
    if dupe_count:
        print(f"   (includes {dupe_count} potential in-file duplicates — DB bizkey sẽ skip)")

    if high:
        print(f"\n🔴 HIGH severity (kiểm tra lại):")
        for w in high[:20]:
            print(f"   {w['detail']}")
        if len(high) > 20:
            print(f"   ... và {len(high) - 20} cảnh báo khác")

    if medium:
        print(f"\n🟡 MEDIUM severity (review nếu cần):")
        for w in medium[:10]:
            print(f"   {w['detail']}")
        if len(medium) > 10:
            print(f"   ... và {len(medium) - 10} cảnh báo khác")

    print(f"\n{'='*60}")
    return warnings


# ── Master data ─────────────────────────────────────────────────

def build_master_data(parsed_rows: list[dict]) -> tuple[dict, dict, dict, dict]:
    """Extract unique master data from parsed rows."""
    sales: dict[str, dict] = {}
    channels: dict[str, dict] = {}
    packages: dict[str, dict] = {}
    customers: dict[str, dict] = {}

    for row in parsed_rows:
        sale_name = row.get("sale_name", "").strip()
        if sale_name and sale_name not in sales:
            sales[sale_name] = {
                "full_name": sale_name,
                "team": row.get("team", ""),
            }

        ch_type = row.get("channel_type_raw", "")
        ch_code = row.get("channel_code", "")
        ch_key = f"{ch_type}|{ch_code}".strip("|")
        if ch_key and ch_key not in channels:
            channels[ch_key] = {
                "channel_code": ch_code or None,
                "name": ch_type or ch_code or None,
                "type": ch_type or None,
            }

        pkg_name = row.get("package_name", "")
        fixed = row.get("fixed", "")
        if pkg_name and pkg_name not in packages:
            packages[pkg_name] = {
                "name": pkg_name,
                "fixed": fixed or None,
            }

        uid = row.get("uid", "")
        if uid and uid not in customers:
            customers[uid] = {
                "uid": uid,
                "full_name": row.get("customer_name") or None,
                "phone": row.get("customer_phone") or None,
                "first_seen": row.get("bank_day") or None,
            }
        elif uid and uid in customers:
            existing_date = customers[uid].get("first_seen")
            new_date = row.get("bank_day")
            if new_date and (not existing_date or new_date < existing_date):
                customers[uid]["first_seen"] = new_date
            if not customers[uid]["full_name"] and row.get("customer_name"):
                customers[uid]["full_name"] = row["customer_name"]
            if not customers[uid]["phone"] and row.get("customer_phone"):
                customers[uid]["phone"] = row["customer_phone"]

    return sales, channels, packages, customers


# ── DB insert ───────────────────────────────────────────────────

def insert_master_data(sb, sales, channels, packages, customers):
    """Insert master data and return lookup dicts (name→id)."""

    print(f"\n👤 Inserting {len(sales)} sales...")
    sale_id_map = {}
    if sales:
        rows = list(sales.values())
        sb.table("sales").upsert(rows, on_conflict="full_name").execute()
        all_sales = sb.table("sales").select("id, full_name").execute()
        sale_id_map = {r["full_name"]: r["id"] for r in (all_sales.data or [])}
    print(f"   → {len(sale_id_map)} sales in DB")

    print(f"\n📺 Inserting {len(channels)} channels...")
    channel_id_map = {}
    if channels:
        for row in channels.values():
            try:
                sb.table("channels").insert(row).execute()
            except Exception:
                pass
        all_ch = sb.table("channels").select("id, name, type, channel_code").execute()
        for r in (all_ch.data or []):
            key = f"{r.get('type') or ''}|{r.get('channel_code') or ''}".strip("|")
            channel_id_map[key] = r["id"]
            if r.get("name"):
                channel_id_map[r["name"]] = r["id"]
    print(f"   → {len(set(channel_id_map.values()))} channels in DB")

    print(f"\n📦 Inserting {len(packages)} packages...")
    package_id_map = {}
    if packages:
        for row in packages.values():
            try:
                sb.table("packages").insert(row).execute()
            except Exception:
                pass
        all_pkg = sb.table("packages").select("id, name").execute()
        package_id_map = {r["name"]: r["id"] for r in (all_pkg.data or [])}
    print(f"   → {len(package_id_map)} packages in DB")

    print(f"\n🧑 Inserting {len(customers)} customers...")
    if customers:
        rows = list(customers.values())
        batch_size = 500
        inserted = 0
        for i in range(0, len(rows), batch_size):
            batch = rows[i : i + batch_size]
            try:
                sb.table("customers").upsert(batch, on_conflict="uid").execute()
                inserted += len(batch)
            except Exception as exc:
                print(f"   ⚠️ Batch {i}-{i+len(batch)} error: {exc}")
        print(f"   → {inserted} customers upserted")

    return sale_id_map, channel_id_map, package_id_map


def insert_payments(sb, parsed_rows, sale_id_map, channel_id_map, package_id_map):
    """Insert payment rows with FK references."""
    print(f"\n💰 Preparing {len(parsed_rows)} payments...")

    payments = []
    skipped = 0
    no_sale = 0

    for row in parsed_rows:
        sale_name = row.get("sale_name", "").strip()
        sale_id = sale_id_map.get(sale_name)
        if not sale_id:
            no_sale += 1
            continue

        uid = row.get("uid")
        pay_time = row.get("pay_time")
        if not uid or not pay_time:
            skipped += 1
            continue

        ch_type = row.get("channel_type_raw", "")
        ch_code = row.get("channel_code", "")
        ch_key = f"{ch_type}|{ch_code}".strip("|")
        channel_id = channel_id_map.get(ch_key) or channel_id_map.get(ch_type)

        pkg_name = row.get("package_name", "")
        package_id = package_id_map.get(pkg_name) if pkg_name else None

        real_vnd = Decimal(str(row.get("real_pay_vnd", 0)))
        gmv_rmb_raw = row.get("gmv_rmb")
        gmv_rmb = Decimal(str(gmv_rmb_raw)) if gmv_rmb_raw else None
        pt = pay_time if isinstance(pay_time, datetime) else datetime.combine(pay_time, datetime.min.time())
        if pt.tzinfo is None:
            pt = pt.replace(tzinfo=timezone.utc)
        gmv_final = compute_gmv_final(pt, real_vnd, gmv_rmb)

        payments.append({
            "uid": uid,
            "pay_time": pt.isoformat(),
            "bank_day": row.get("bank_day"),
            "package_id": package_id,
            "sale_id": sale_id,
            "channel_id": channel_id,
            "real_pay_vnd": float(real_vnd),
            "gmv_rmb": float(gmv_rmb) if gmv_rmb else None,
            "gmv_final": float(gmv_final),
            "payment_seq": row.get("payment_seq"),
            "team": row.get("team", ""),
            "note": row.get("note"),
            "status": "active",
        })

    print(f"   Ready: {len(payments)} | No sale: {no_sale} | Skipped: {skipped}")
    print(f"\n   Inserting...")

    batch_size = 100
    inserted = 0
    dupes = 0
    errors = 0

    for i in range(0, len(payments), batch_size):
        batch = payments[i : i + batch_size]
        try:
            res = sb.table("payments").insert(batch).execute()
            inserted += len(res.data or batch)
        except Exception as exc:
            err_str = str(exc).lower()
            if "duplicate" in err_str or "payments_bizkey" in err_str:
                for p in batch:
                    try:
                        sb.table("payments").insert(p).execute()
                        inserted += 1
                    except Exception:
                        dupes += 1
            else:
                print(f"   ⚠️ Batch {i} error: {exc}")
                errors += len(batch)

        progress = min(i + batch_size, len(payments))
        print(f"   ... {progress}/{len(payments)}")

    print(f"\n📊 Insert results:")
    print(f"   Inserted:  {inserted}")
    print(f"   Dupes:     {dupes}")
    print(f"   No sale:   {no_sale}")
    print(f"   Skipped:   {skipped}")
    print(f"   Errors:    {errors}")


# ── Main ────────────────────────────────────────────────────────

def main():
    if len(sys.argv) < 2:
        print("Usage: python migrate_gsheet.py <path-to-xlsx> [--commit]")
        print()
        print("  Default:   dry-run (parse + validate, không ghi DB)")
        print("  --commit:  chạy thật (parse + validate + ghi DB)")
        print()
        print("Sheet tab names are auto-matched to parsers (HCM/Danang/SM Hanoi).")
        print("If a tab isn't recognized, edit SHEET_PARSERS at the top of this file.")
        sys.exit(1)

    xlsx_path = sys.argv[1]
    commit = "--commit" in sys.argv

    if not os.path.exists(xlsx_path):
        print(f"❌ File not found: {xlsx_path}")
        sys.exit(1)

    url = os.getenv("SUPABASE_URL", "").strip()
    key = os.getenv("SUPABASE_SERVICE_ROLE_KEY", "").strip()
    if not url or not key:
        print("❌ Missing SUPABASE_URL or SUPABASE_SERVICE_ROLE_KEY in .env")
        sys.exit(1)

    mode = "🔴 COMMIT (ghi DB)" if commit else "🟢 DRY-RUN (chỉ validate)"
    print(f"\n{'='*60}")
    print(f"   Migration mode: {mode}")
    print(f"{'='*60}")

    # Step 1: Parse
    parsed_rows = read_all_sheets(xlsx_path)
    if not parsed_rows:
        print("❌ No data parsed. Check sheet tab names vs SHEET_PARSERS mapping.")
        sys.exit(1)

    # Step 2: Validate
    warnings = validate_rows(parsed_rows)

    # Step 3: Master data summary
    sales, channels, packages, customers = build_master_data(parsed_rows)
    print(f"\n📋 Master data found:")
    print(f"   Sales:     {len(sales)}")
    print(f"   Channels:  {len(channels)}")
    print(f"   Packages:  {len(packages)}")
    print(f"   Customers: {len(customers)}")

    if not commit:
        high = [w for w in warnings if w["severity"] == "high"]
        print(f"\n🟢 DRY-RUN complete. {len(parsed_rows)} rows parsed, {len(high)} high warnings.")
        print(f"   Review output above. Khi OK, chạy lại với --commit:")
        print(f'   python migrate_gsheet.py "{xlsx_path}" --commit')
        return

    # Step 4: Confirm
    high = [w for w in warnings if w["severity"] == "high"]
    if high:
        print(f"\n⚠️  Có {len(high)} cảnh báo HIGH. Tiếp tục insert? (y/n): ", end="")
        if input().strip().lower() != "y":
            print("❌ Aborted.")
            return

    # Step 5: Insert
    sb = create_client(url, key)

    sale_id_map, channel_id_map, package_id_map = insert_master_data(
        sb, sales, channels, packages, customers
    )

    insert_payments(sb, parsed_rows, sale_id_map, channel_id_map, package_id_map)

    print("\n✅ Migration complete!")


if __name__ == "__main__":
    main()
